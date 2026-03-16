import pandas as pd
import json
from datetime import datetime
from openpyxl.styles import Font, Alignment

# --- SEGÉDFÜGGVÉNYEK ---

def format_project_status(status):
    status_map = {"ongoing": "Folyamatban", "completed": "Befejezve", "cancelled": "Megszakítva"}
    return status_map.get(status, status)

def get_hours(start_str, end_str, break_min):
    try:
        start = datetime.fromisoformat(start_str.replace('Z', '+00:00'))
        end = datetime.fromisoformat(end_str.replace('Z', '+00:00'))
        delta = end - start
        return round((delta.total_seconds() / 3600) - (break_min / 60), 2)
    except: return 0

# --- FORMÁZÓ FÜGGVÉNYEK ---

def format_project_sheet(ws):
    bold_font = Font(bold=True)
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    for row_idx in [2, 3, 9]:
        for col_idx in range(2, 5):
            ws.cell(row=row_idx, column=col_idx).font = bold_font

def format_materials_sheet(ws, data_len, total_sum):
    bold_font = Font(bold=True)
    red_bold = Font(bold=True, color="FF0000")
    widths = {'B': 15, 'C': 35, 'D': 12, 'E': 10, 'F': 15, 'G': 15, 'H': 20}
    for col, width in widths.items(): ws.column_dimensions[col].width = width
    for cell in ws[2]: 
        if cell.column >= 2: cell.font = bold_font
    
    summary_row = data_len + 3
    ws.cell(row=summary_row, column=2, value="ÖSSZESEN").font = red_bold
    total_cell = ws.cell(row=summary_row, column=8, value=total_sum)
    total_cell.font = red_bold
    total_cell.number_format = '#,##0'

def format_worklog_sheet(ws, data_len, total_hours, total_money):
    bold_font = Font(bold=True)
    red_bold = Font(bold=True, color="FF0000")
    widths = {'B': 12, 'C': 20, 'D': 15, 'E': 15, 'F': 25, 'G': 8, 'H': 12, 'I': 15, 'J': 12, 'K': 15}
    for col, width in widths.items(): ws.column_dimensions[col].width = width
    for cell in ws[2]:
        if cell.column >= 2: cell.font = bold_font
    
    summary_row = data_len + 3
    ws.cell(row=summary_row, column=2, value="MINDÖSSZESEN").font = red_bold
    ws.cell(row=summary_row, column=7, value=total_hours).font = red_bold
    ws.cell(row=summary_row, column=9, value=total_money).font = red_bold

def format_machines_sheet(ws, data_len, machine_totals):
    bold_font = Font(bold=True)
    red_bold = Font(bold=True, color="FF0000")
    widths = {'B': 15, 'C': 25, 'D': 15, 'E': 15, 'F': 15}
    for col, width in widths.items(): ws.column_dimensions[col].width = width
    for cell in ws[2]:
        if cell.column >= 2: cell.font = bold_font

    # Gép alapú összesítők hozzáadása a táblázat alá
    current_row = data_len + 3
    for m_name, total in machine_totals.items():
        ws.cell(row=current_row, column=2, value="ÖSSZESEN").font = red_bold
        ws.cell(row=current_row, column=3, value=m_name).font = red_bold
        ws.cell(row=current_row, column=6, value=total).font = red_bold
        current_row += 1

# --- FŐ PROGRAM ---

with open("test.json", encoding="utf-8") as f:
    data = json.load(f)

# 1. Projekt & Alapanyagok feldolgozása (Változatlan)
p = data["project"]
project_content = [
    ["A (Megnevezés)", "B (Adat / Érték)", "C (Mértékegység)"],
    ["PROJEKT ADATOK", None, None],
    ["Projekt Neve", p["projectName"], None],
    ["Megrendelő", p["customerName"], None],
    ["Helyszín", p["projectLocation"], None],
    ["Státusz", format_project_status(p["projectStatus"]), None],
    [None, None, None],
    ["KAPCSOLAT", "(Ügyfél adatok)", None],
    ["Email", p["customerEmail"], None],
    ["Telefon", p["customerPhone"], None],
]

mat_rows, total_mat_sum = [], 0
for m in data.get("material", []):
    price_mode = "Egységár" if m.get("priceMode") == "unitPrice" else "Egyedi ár"
    mat_rows.append([m["date"][:10].replace("-", "."), m["name"], m["quantity"], m["unit"], m.get("unitPrice", ""), price_mode, m["price"]])
    total_mat_sum += m["price"]

# 2. Munkadíjak feldolgozása
roles_map = {1: "Admin", 2: "Építésvezető", 3: "Kertész"}
users_dict = {u["id"]: u for u in data.get("users", [])}
work_rows, total_work_hours, total_work_money = [], 0, 0

for log in data.get("worklog", []):
    user = users_dict.get(log["employeeId"], {})
    hours = get_hours(log["startTime"], log["endTime"], log.get("breakMinutes", 0))
    salary = user.get("salary") or 0
    work_rows.append([log["date"][:10].replace("-", "."), user.get("name", "Ismeretlen"), roles_map.get(user.get("role"), ""), "", log.get("description", "").replace("\n", " "), hours, salary, hours*salary, salary, hours*salary])
    total_work_hours += hours
    total_work_money += hours * salary

# 3. Munkagépek feldolgozása
machines_lookup = {m["id"]: m["name"] for m in data.get("machines", [])}
machine_rows = []
machine_totals = {}

for mlog in data.get("machineWorklog", []):
    m_name = machines_lookup.get(mlog["machineId"], "Ismeretlen gép")
    daily_usage = float(mlog["newHours"]) - float(mlog["previousHours"])
    machine_rows.append([
        mlog["date"][:10].replace("-", "."),
        m_name,
        mlog["previousHours"],
        mlog["newHours"],
        daily_usage
    ])
    machine_totals[m_name] = machine_totals.get(m_name, 0) + daily_usage

# 4. Mentés
file_name = "projekt_jelentes_vegleges.xlsx"
df_mat = pd.DataFrame(mat_rows, columns=["Dátum", "Anyag Megnevezése", "Mennyiség", "Egység", "Egységár (Ft)", "Ár Módja", "VÉGÖSSZEG (Ft)"])
df_mat['Dátum'] = pd.to_datetime(df_mat['Dátum'])
df_mat = df_mat.sort_values(by='Dátum')
df_mat['Dátum'] = df_mat['Dátum'].dt.strftime('%Y.%m.%d.')

# 2. Munkadíjak rendezése
df_work = pd.DataFrame(work_rows, columns=["Dátum", "Dolgozó Neve", "Szerepkör", "Munka Típusa", "Leírás", "Óra", "Díj (Ft/óra)", "Összesen (Ft)", "Órabér (Ft)", "ÖSSZESEN (Ft)"])
df_work['Dátum'] = pd.to_datetime(df_work['Dátum'])
df_work = df_work.sort_values(by='Dátum')
df_work['Dátum'] = df_work['Dátum'].dt.strftime('%Y.%m.%d.')

# 3. Munkagépek rendezése
df_mach = pd.DataFrame(machine_rows, columns=["Dátum", "Gép Neve", "Kezdő Óraállás", "Záró Óraállás", "Napi Üzemóra"])
df_mach['Dátum'] = pd.to_datetime(df_mach['Dátum'])
df_mach = df_mach.sort_values(by='Dátum')
df_mach['Dátum'] = df_mach['Dátum'].dt.strftime('%Y.%m.%d.')

# --- Mentés az ExcelWriter-rel ---
with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
    # Most már a rendezett DataFrame-eket írjuk ki:
    pd.DataFrame(project_content).to_excel(writer, sheet_name='Projekt', startrow=1, startcol=1, index=False, header=False)
    df_mat.to_excel(writer, sheet_name='Alapanyagok', startrow=1, startcol=1, index=False)
    df_work.to_excel(writer, sheet_name='Munkadíjak', startrow=1, startcol=1, index=False)
    df_mach.to_excel(writer, sheet_name='Munkagépek', startrow=1, startcol=1, index=False)
    
    # Formázó függvények hívása (ws paraméterrel)
    format_project_sheet(writer.sheets['Projekt'])
    format_materials_sheet(writer.sheets['Alapanyagok'], len(df_mat), total_mat_sum)
    format_worklog_sheet(writer.sheets['Munkadíjak'], len(df_work), total_work_hours, total_work_money)
    format_machines_sheet(writer.sheets['Munkagépek'], len(df_mach), machine_totals)

print(f"Sikeres mentés: {file_name}")